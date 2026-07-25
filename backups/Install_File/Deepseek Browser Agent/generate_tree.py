import os
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Configuration
ROOT_DIR = Path(r"E:\BOT_FINALBOT13 STG\BOT_FINALBOT")
OUTPUT_FILE = Path(r"E:\BOT_FINALBOT13 STG\BOT_FINALBOT\docs\Payload_Audit_Report 2.xlsx")

# Excluded directories
EXCLUDED_DIRS = {
    '.venv', '.git', 'node_modules', '__pycache__', 
    '.agents', '.pytest_cache', 'browser_profile', 'backups', 'logs'
}

# Excluded files
EXCLUDED_FILES = {'__init__.py'}

# Excluded folder names (partial matches)
EXCLUDED_FOLDER_NAMES = {'Deepseek Browser Agent'}

def should_exclude(path):
    """Check if path should be excluded from scanning"""
    path_parts = path.parts
    
    # Check if any part of the path matches excluded directories
    for part in path_parts:
        if part in EXCLUDED_DIRS:
            return True
        # Check for partial matches in folder names
        for excluded_name in EXCLUDED_FOLDER_NAMES:
            if excluded_name in part:
                return True
    
    # Check if file name is excluded
    if path.is_file() and path.name in EXCLUDED_FILES:
        return True
    
    return False

def get_tree_structure(root_path, prefix="", is_last=True, level=0):
    """Generate tree structure with symbols"""
    lines = []
    
    # Get all items in directory, excluding those we don't want
    try:
        items = sorted([p for p in root_path.iterdir() if not should_exclude(p)], 
                      key=lambda x: (not x.is_dir(), x.name.lower()))
    except PermissionError:
        return lines
    
    for i, item in enumerate(items):
        is_last_item = (i == len(items) - 1)
        connector = "└── " if is_last_item else "├── "
        
        # Build the line with proper indentation
        line = prefix + connector + item.name
        if item.is_dir():
            line += "/"
        lines.append(line)
        
        # Recursively process directories
        if item.is_dir():
            extension = "    " if is_last_item else "│   "
            lines.extend(get_tree_structure(item, prefix + extension, is_last_item, level + 1))
    
    return lines

def generate_excel_report(tree_lines, output_path):
    """Generate Excel report with styled tree"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "File Tree"
    
    # Set column widths
    ws.column_dimensions['A'].width = 100
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    
    # Style for header
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="left", vertical="center")
    
    # Headers
    headers = ["File Tree Structure", "Type", "Level"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Populate tree data
    for idx, line in enumerate(tree_lines, start=2):
        # Determine type and level
        is_dir = line.endswith("/")
        item_type = "Folder" if is_dir else "File"
        level = (len(line) - len(line.lstrip("│├└─ ")))
        
        # Clean up the line for display
        display_line = line
        
        # Apply color coding based on type
        cell = ws.cell(row=idx, column=1, value=display_line)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        if is_dir:
            cell.font = Font(bold=True, color="1F4E79")
        else:
            cell.font = Font(color="333333")
        
        ws.cell(row=idx, column=2, value=item_type)
        ws.cell(row=idx, column=3, value=level)
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # Summary data
    total_files = sum(1 for line in tree_lines if not line.endswith("/"))
    total_folders = sum(1 for line in tree_lines if line.endswith("/"))
    
    summary_data = [
        ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Root Directory", str(ROOT_DIR)],
        ["Total Files", total_files],
        ["Total Folders", total_folders],
        ["Total Items", total_files + total_folders],
    ]
    
    for row, (key, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=value)
    
    # Save the workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    return total_files, total_folders

def main():
    """Main execution"""
    print(f"Scanning directory: {ROOT_DIR}")
    print(f"Excluded: {', '.join(EXCLUDED_DIRS)}, {', '.join(EXCLUDED_FILES)}, Deepseek Browser Agent")
    
    # Generate tree structure
    tree_lines = get_tree_structure(ROOT_DIR)
    
    if not tree_lines:
        print("No items found or directory is empty.")
        return
    
    # Generate Excel report
    total_files, total_folders = generate_excel_report(tree_lines, OUTPUT_FILE)
    
    print(f"\n✅ Tree structure generated successfully!")
    print(f"📊 Output: {OUTPUT_FILE}")
    print(f"📁 Total folders: {total_folders}")
    print(f"📄 Total files: {total_files}")
    print(f"📦 Total items: {total_folders + total_files}")

if __name__ == "__main__":
    main()
